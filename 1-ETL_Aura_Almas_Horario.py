import requests
import pandas as pd
import urllib3
import os
import logging

from sqlalchemy import create_engine
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from supabase import create_client, Client
from supabase import create_client, Client
from io import BytesIO
from urllib.parse import urlparse, parse_qs

# =======================================
# Consultando o banco de dados SQL da F2M
# =======================================
# função para coleta dos dados
def consulta_dados_transporte():
    server = 'sql-f2m-databases-2-prod.miningcontrol.cloud'
    database = 'mining_control_ama'
    username = 'aura_almas_bi'
    password = 'c3YFnEL9BHJzRrEDXTtYm'

    # String de conexão compatível com SQLAlchemy
    conn_str = (
        f"mssql+pyodbc://{username}:{password}@{server}/{database}"
        "?driver=ODBC+Driver+18+for+SQL+Server"
        "&Encrypt=yes"
        "&TrustServerCertificate=yes"
    )

    # Criação da engine SQLAlchemy
    engine = create_engine(conn_str)

#Coleta os dados dos ultimos 33 dias para exibição dos dados acumulados do mês
    end_time = datetime.now()
    start_time = (end_time - timedelta(days=33)).replace(hour=0, minute=0, second=0, microsecond=0)
    start_str = start_time.strftime('%Y-%m-%d %H:%M:%S')
    end_str = end_time.strftime('%Y-%m-%d %H:%M:%S')

    colunas_selecionadas = ['datetime_end','origin','destination_subarea','exception_type','material_group', 'calculated_mass','material']
    colunas_sql = ', '.join(colunas_selecionadas)

    query = f"""
    SELECT {colunas_sql}
    FROM [dbo].[dw_transport_report]
    WHERE [datetime_end] >= '{start_str}' AND [datetime_end] <= '{end_str}'
    """

    try:
        df = pd.read_sql(query, engine)
    except Exception as e:
        print("Erro ao executar a consulta SQL com SQLAlchemy:", e)
        df = pd.DataFrame()

    return df

# Data frame final sem ajustes
df_transporte = consulta_dados_transporte()

#===========================================================
# Coleta de dados do PI Web API - Metodo de Agregação Total
#===========================================================

# Desativa o aviso de conexão insegura
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Função para arredondar segundos para múltiplos de 0 ou 5
def round_to_last_0_or_5_seconds(dt):
    seconds = dt.second
    if seconds % 10 < 5:
        rounded_seconds = seconds - (seconds % 10)
    else:
        rounded_seconds = seconds - (seconds % 10) + 5
    rounded = dt - timedelta(seconds=seconds - rounded_seconds, microseconds=dt.microsecond)
    return rounded

# Função para consultar valores agregados (Total por hora) de um WebID
def fetch_data_for_webid(webid, start_time, end_time, interval):
    base_url = "https://10.135.7.10/piwebapi/streams/"
    url = f"{base_url}{webid}/summary"

    start_time = round_to_last_0_or_5_seconds(start_time)
    end_time = round_to_last_0_or_5_seconds(end_time)

    params = {
        "startTime": start_time.isoformat(),
        "endTime": end_time.isoformat(),
        "summaryType": "Total",
        "summaryDuration": interval,  # Ex: "1h"
        "calculationBasis": "TimeWeighted",
        "timeType": "Auto"
    }

    response = requests.get(url, headers=headers, params=params, verify=False)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Erro ao obter dados para WebID {webid}: {response.status_code} - {response.text}")
        return None

# Cabeçalhos da requisição
headers = {
    "Accept": "application/json",
    "Authorization": "Basic dGFyY2l6by5qdW5pb3I6QW5uZUBhdXJhMjAyNQ=="  # Insira o token se necessário
}

# Caminho do Excel com WebIDs e Apelidos
file_path = r'C:\Users\tarcizo.junior\OneDrive - Aura Minerals\04 - Projetos\Python\Dash_Aura_Almas_Horario\Dados\dados_piwebapi.xlsx'
piwebapi_data = pd.read_excel(file_path)

# Lista de IDs desejados
ids = [11]  # Substitua conforme necessário
webids = piwebapi_data[piwebapi_data['Id'].isin(ids)]['WebId'].tolist()

# Datas de início e fim
end_date = datetime.today().date() + timedelta(days=1)
start_date = end_date - timedelta(days=33)

start_time = datetime.fromisoformat(start_date.strftime('%Y-%m-%d'))
end_time = datetime.fromisoformat(end_date.strftime('%Y-%m-%d'))

interval = "1h"  # Intervalo de agregação

# DataFrame final
df_totalizador = pd.DataFrame()
webids_com_erro = []

# Coleta dos dados
for webid in webids:
    try:
        data = fetch_data_for_webid(webid, start_time, end_time, interval)
        if data:
            items = data.get('Items', [])
            if items:
                good_items = [item['Value'] for item in items if item.get('Value') and item['Value'].get('Good') and item['Value'].get('Value') is not None]

                if good_items:
                    df = pd.DataFrame(good_items)

                    if 'Timestamp' in df.columns and 'Value' in df.columns:
                        # Multiplica os valores por 24
                        df['Value'] = (df['Value'] * 24*0.96).round(2)

                        # Ajusta fuso horário
                        df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.tz_convert('Etc/GMT+3').dt.strftime('%Y-%m-%d %H:%M:%S')

                        # Renomeia a coluna com o apelido
                        column_name = piwebapi_data.loc[piwebapi_data['WebId'] == webid, 'Apelido'].values[0]
                        df.rename(columns={'Value': column_name}, inplace=True)

                        # Mescla no DataFrame final
                        if df_totalizador.empty:
                            df_totalizador = df
                        else:
                            df_totalizador = pd.merge(df_totalizador, df, on='Timestamp', how='outer')
                    else:
                        print(f"As colunas 'Timestamp' e 'Value' não estão presentes para WebID {webid}.")
                else:
                    print(f"Não há itens válidos (Good=True) para WebID {webid}.")
            else:
                print(f"Não há itens retornados para WebID {webid}.")
    except Exception as e:
        print(f"Erro ao processar WebID {webid}: {e}")
        webids_com_erro.append(webid)

#====================================================
# Coleta de dados do PI Web API - Metodo Interpolação
#====================================================

# Lista de IDs de variáveis de vazão
ids_vazao = [11]  # Substitua pelos IDs desejados

# DataFrame final para dados de vazão
df_vazao_final = pd.DataFrame()

# Intervalo de coleta em segundos
intervalo_vazao = "30s"

# Loop para cada ID de vazão
for id_vazao in ids_vazao:
    try:
        # Buscar WebID e Apelido da planilha
        webid_vazao = piwebapi_data.loc[piwebapi_data['Id'] == id_vazao, 'WebId'].values[0]
        apelido_vazao = piwebapi_data.loc[piwebapi_data['Id'] == id_vazao, 'Apelido'].values[0]

        # Construir URL de interpolated values
        url_vazao = f"https://10.135.7.10/piwebapi/streams/{webid_vazao}/interpolated"

        # Ajustar datas para interpolated (considera os daddos das variaveis declaradas no inicio do bloco)
        start_time_vazao = round_to_last_0_or_5_seconds(start_time)
        end_time_vazao = round_to_last_0_or_5_seconds(end_time)

        params_vazao = {
            "startTime": start_time_vazao.isoformat(),
            "endTime": end_time_vazao.isoformat(),
            "interval": intervalo_vazao
        }

        response = requests.get(url_vazao, headers=headers, params=params_vazao, verify=False)

        if response.status_code == 200:
            data = response.json()
            items = data.get('Items', [])
            good_items = [item for item in items if item.get('Good') and item.get('Value') is not None]

            if good_items:
                df = pd.DataFrame(good_items)
                df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.tz_convert('Etc/GMT+3')

                # Calcular a média móvel de 200 segundos (aproximadamente 3-4 pontos para dados a cada 60s)
                df.set_index('Timestamp', inplace=True)
                df[apelido_vazao + "_mm200s"] = df['Value'].rolling('200s').mean()
                df.reset_index(inplace=True)

                # Manter apenas timestamp e a média móvel
                df = df[['Timestamp', apelido_vazao + "_mm200s"]]

                # Unir ao dataframe final de vazões
                if df_vazao_final.empty:
                    df_vazao_final = df
                else:
                    df_vazao_final = pd.merge(df_vazao_final, df, on='Timestamp', how='outer')
            else:
                print(f"Sem dados válidos para ID {id_vazao}")
        else:
            print(f"Erro na requisição para WebID de ID {id_vazao}: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Erro ao processar ID {id_vazao}: {e}")

# ========================================================
# Coleta dos dados do Onedrive empresarial (link publico)
# ========================================================

#Função para baixar e criar dataframe a partir da leitura do arquivo de excel
def baixar_arquivos_onedrive(link: str, sheet_name: str) -> pd.DataFrame | None:
    try:
        if "?download=1" not in link:
            link += "?download=1"

        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(link, headers=headers)
        # Verifica se é HTML (erro comum)
        content_type = response.headers.get("Content-Type", "")
        if "text/html" in content_type.lower():
            print(f"❌ Conteúdo baixado é uma página HTML, não um arquivo Excel. Verifique o link: {link}")
            return None
        if response.status_code != 200:
            print(f"❌ Erro ao baixar o arquivo: {response.status_code} | URL: {link}")
            return None
        # Lê o conteúdo como Excel
        file = BytesIO(response.content)
        df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
        return df
    except Exception as e:
        print(f"❌ Erro ao processar o arquivo: {e}")
        return None

# Link de compartilhamento dos arquivos
arquivo_dia_atual = "https://auraminerals-my.sharepoint.com/:x:/p/tarcizo_junior/EdD7gQG1UsFCtkZu3JNgInsB9o6EE5TGY3vNRA41IKCZZQ?download=1" #codigos gerado no onedrive busines (pessoal corporativo)
arquivo_historico = "https://auraminerals-my.sharepoint.com/:x:/p/tarcizo_junior/EWpS3k8SQTtHmP9Kp42u9ykBAWIZFie3z9tH8G31WOxBCw?download=1" #codigos gerado no onedrive busines (pessoal corporativo)

# Carregamento das def's
dados_planta_dia_atual = baixar_arquivos_onedrive(arquivo_dia_atual, sheet_name="Dados_painel_hora_hora")
dados_planta_historico = baixar_arquivos_onedrive(arquivo_historico, sheet_name="BancoDadosVariáveis")

# ============================================================================================
# Função de filtro de período (aplicar diretamente no dataframe, não aplicavel nas consultas)
# ============================================================================================

#Função garante que os dados estarão no periodo de tempo desejado (ultimas 24h excluindo a hora atual)
def filtrar_por_intervalo_de_tempo(df, coluna_data, inicio, fim):
    # Converte a coluna para datetime, se não for
    df[coluna_data] = pd.to_datetime(df[coluna_data], errors='coerce')
    # Aplica o filtro
    df_filtrado = df[(df[coluna_data] >= inicio) & (df[coluna_data] < fim)]
    return df_filtrado

# Definir intervalo de tempo dinamicamente
agora = datetime.now().replace(minute=0, second=0, microsecond=0)
inicio = agora - timedelta(hours=24)

# ============================
# Tratamento dos dados de mina
# ============================

# Garante que as colunas estão como string e padroniza valores
df_transporte['origin'] = df_transporte['origin'].astype(str).str.strip()
df_transporte['destination_subarea'] = df_transporte['destination_subarea'].astype(str).str.strip().str.lower() #precisa converter para letra minuscula para garantir que o filtro irá funcionar corretamente
df_transporte['exception_type'] = df_transporte['exception_type'].astype(str).str.strip().str.lower() #precisa converter para letra minuscula para garantir que o filtro irá funcionar corretamente
df_transporte['calculated_mass'] = pd.to_numeric(df_transporte['calculated_mass'],errors='coerce').astype('float64') #garantir que a coluna esta como numero decimal

# Filtros aplicados um a um
df_temp = df_transporte[df_transporte['origin'].isin(['Cava Paiol', 'Cava Sul'])]
df_temp = df_temp[df_temp['destination_subarea'] != 'acesso dentro da cava']
df_transporte_filtrado = df_temp[df_temp['exception_type'] != 'Edited_Delete'].copy()

# Arredondar as horas com minutos para hora exata
df_transporte_filtrado.loc[:, 'hora_completa'] = pd.to_datetime(df_transporte_filtrado['datetime_end']).dt.floor('h')

#filtrar intervalo de tempo (ultimas 24h excluindo a hora atual)
#df_transporte_filtrado = filtrar_por_intervalo_de_tempo(df_transporte_filtrado, 'datetime_end', inicio, agora) # esta desativado esse filtro esta aplicado no dash

#Ajustar litologias no dataframe
df_transporte_filtrado['material'] = df_transporte_filtrado['material'].replace({
    'Estéril': 'Estéril',
    'Estéril-RI': 'Estéril',
    'HG': 'HG',
    'HG1': 'HG',
    'HG2': 'HG',
    'HG3': 'HG',
    'LG': 'LG',
    'LG1': 'LG',
    'LG2': 'LG',
    'LG3': 'LG',
    'LG4': 'LG',
    'LG5': 'LG',
    'MG': 'MG',
    'MW': 'MG'
})

# =========================================================
# Tratamento dos dados de Planta via API (Massa Alimentada)
# =========================================================
# Excluir colunas desnecessarias
df_totalizador['Timestamp'] = pd.to_datetime(df_totalizador['Timestamp'], errors='coerce')
df_totalizador = df_totalizador.drop(columns=['UnitsAbbreviation','Good','Questionable','Substituted','Annotated'])
#filtrar intervalo de tempo (ultimas 24h excluindo a hora atual)
#df_totalizador = filtrar_por_intervalo_de_tempo(df_totalizador, 'Timestamp', inicio, agora) # esta desativado esse filtro esta aplicado no dash
#converter coluna timestamp para string
df_totalizador['Timestamp'] = df_totalizador['Timestamp'].dt.strftime('%Y-%m-%dT%H:%M:%S')
#Mudar o nome da coluna para garantir compatibilidade na hora de unir os Dataframes
df_totalizador.rename(columns={"Retomada - TR02 - Balança": "Moinho_Massa Alimentada Moagem_(t)"},inplace=True)

# =============================================================
# Tratamento dos dados de Planta via API (Vazão de Alimentação)
# =============================================================
#Eliminar formato date time zone
df_vazao_final['Timestamp'] = df_vazao_final['Timestamp'].dt.tz_localize(None)
# cRiar coluna com data e hora somente
df_vazao_final.loc[:,'data_hora_exata'] = pd.to_datetime(df_vazao_final['Timestamp']).dt.floor('h')
#filtrar intervalo de tempo (ultimas 24h excluindo a hora atual)
#df_vazao_final = filtrar_por_intervalo_de_tempo(df_vazao_final, 'Timestamp', inicio, agora) # esta desativado esse filtro esta aplicado no dash

# ===============================================================
# Tratamento dos dados da planta via Onedrive (Dados Historicos)
# ===============================================================

#Eliminar colunas desnecessarias
dados_planta_historico = dados_planta_historico.loc[:, [
    "Data",
    "Hora corrigida",
    "Britagem_Massa Produzida Britagem_(t)",
    "Britagem_Justificativa de NÂO atingir a massa_(txt)",
    "Moinho_Justificativa de NÂO atingir a massa_(txt)",
    "Moinho_Justificativa do Tempo operando com taxa a menor_(txt)"
]]

# Converte a coluna "Hora corrigida" para datetime, só horário mesmo
dados_planta_historico["Hora corrigida"] = pd.to_datetime(
    dados_planta_historico["Hora corrigida"], errors="coerce", format="%H:%M:%S"
)

# Filtrar somente dados necessarios (filtrando os dados dos ultimo 33 dias para acumulação da produção do mês atual)
# Converter a coluna 'Data' para datetime (se ainda não for)
dados_planta_historico["Data"] = pd.to_datetime(dados_planta_historico["Data"], errors="coerce", dayfirst=True)

# Data de hoje (sem hora)
hoje = datetime.now().date()

# Dois dias anteriores (sem hora)
data_inicio = hoje - timedelta(days=33)
data_fim = hoje - timedelta(days=1)

# Filtra os registros
dados_planta_historico = dados_planta_historico[
    (dados_planta_historico["Data"].dt.date >= data_inicio) &
    (dados_planta_historico["Data"].dt.date <= data_fim)
]

# Converte a coluna "Hora corrigida" para datetime
dados_planta_historico["Hora corrigida"] = pd.to_datetime(
    dados_planta_historico["Hora corrigida"], errors="coerce", format="%H:%M:%S"
)
# Cria 'Timestamp' juntando Data + Hora (sem criar coluna extra) - coluna será utilizada nas proximas etapas
dados_planta_historico["Timestamp"] = pd.to_datetime(
    dados_planta_historico["Data"].dt.strftime("%Y-%m-%d") + " " +
    dados_planta_historico["Hora corrigida"].dt.strftime("%H:%M:%S"),
    format="%Y-%m-%d %H:%M:%S",
    errors="coerce"
)

# ========================================================
# Tratamento dos dados da planta via Onedrive (Dia Atual)
# ========================================================

# Garante que a coluna 'Data' esteja no formato datetime
dados_planta_dia_atual["Data"] = pd.to_datetime(
    dados_planta_dia_atual["Data"], errors="coerce", dayfirst=True
)

# Data atual (sem hora)
hoje = datetime.now().date()

# Filtra os registros que são do dia atual - aplicado para evitar que se a sala de controle demore a trocar de dia no relatorio não apareça dados duplicados em df_dados_planta
dados_planta_dia_atual = dados_planta_dia_atual[
    dados_planta_dia_atual["Data"].dt.date == hoje
]

# Converte 'Data' para datetime
dados_planta_dia_atual["Data"] = pd.to_datetime(
    dados_planta_dia_atual["Data"], errors="coerce", dayfirst=True
)

# Cria 'Timestamp' juntando Data + Hora (sem criar coluna extra)
dados_planta_dia_atual["Timestamp"] = pd.to_datetime(
    dados_planta_dia_atual["Data"].dt.strftime("%Y-%m-%d") + " " +
    dados_planta_dia_atual["Hora corrigida"].dt.strftime("%H:%M:%S"),
    format="%Y-%m-%d %H:%M:%S",
    errors="coerce"
)

# ===========================================================================
# Criação do Dataframe final com os dados da planta consolidado (PI + Excel)
# ===========================================================================

# União dos dataframes
df_dados_planta = pd.concat([dados_planta_historico, dados_planta_dia_atual], ignore_index=True)

# Filtrar dados no intervalo de tempo das ultimas 24h excluindo a hora atual
#df_dados_planta = filtrar_por_intervalo_de_tempo(df_dados_planta, 'Timestamp', inicio, agora) # Desativado esta sendo aplicado no das

# Garantir que ambas as colunas estejam em datetime completo (com hora)
df_dados_planta["Timestamp"] = pd.to_datetime(df_dados_planta["Timestamp"], errors="coerce")
df_totalizador["Timestamp"] = pd.to_datetime(df_totalizador["Timestamp"], errors="coerce")

# Fazer o merge usando 'Timestamp' como chave
df_dados_planta = df_totalizador.merge(
    df_dados_planta[[
        "Timestamp",
        "Britagem_Massa Produzida Britagem_(t)",
        "Britagem_Justificativa de NÂO atingir a massa_(txt)",
        "Moinho_Justificativa de NÂO atingir a massa_(txt)",
        "Moinho_Justificativa do Tempo operando com taxa a menor_(txt)"]],
    on="Timestamp",
    how="left"  
)

#Exclusão de colunas desnecessarias
#df_dados_planta=df_dados_planta.drop(columns=['Hora corrigida','Data'])

# ==========================
# Exportar dados para .xlsx
# ==========================

# Caminho de destino
destinos_arquivos = r'C:\Users\Dataminds\Aura Minerals\Almas - Performance - Data Minds - Data Minds\03 - Diretoria\Painel Hora a Hora'

# === Função de exportação adaptada ===
def salvar_dataframes_como_tabelas_padrao(dfs: dict, diretorio_destino: str):
    caminho_saida = os.path.join(diretorio_destino, 'Dados_Teste.xlsx')

    # Salva os DataFrames em planilhas individuais
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        for nome_df, df in dfs.items():
            df.to_excel(writer, sheet_name=nome_df, index=False)

    # Abre o arquivo para aplicar formatações com openpyxl
    wb = load_workbook(caminho_saida)

    for nome_df in dfs.keys():
        ws = wb[nome_df]
        ws.sheet_view.showGridLines = False

        max_row = ws.max_row
        max_col = ws.max_column
        letra_final = get_column_letter(max_col)
        ref_tabela = f"A1:{letra_final}{max_row}"

        nome_tabela = f"TB_{nome_df}".lower().replace(" ", "_")[:25]

        tabela = Table(displayName=nome_tabela, ref=ref_tabela)
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

        # Ajuste da largura das colunas
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=max_row, max_col=max_col), start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(caminho_saida)
    logging.info(f"Arquivo 'Dados_Transporte_Mina.xlsx' salvo com sucesso em: {caminho_saida}")

# Chamada da função com os novos DataFrames
salvar_dataframes_como_tabelas_padrao(
    dfs={
        'df_alimentacao_moagem': df_totalizador,
        'df_media_movel': df_vazao_final,
        'df_transporte_filtrado': df_transporte_filtrado,
        'dados_planta': df_dados_planta
    },
    diretorio_destino=destinos_arquivos
)

# ================================================================
# Conversão de dados para garantir compatibilidade com o supabase
# ================================================================
#converter coluna timestamp para string
df_transporte_filtrado['datetime_end'] = df_transporte_filtrado['datetime_end'].dt.strftime('%Y-%m-%dT%H:%M:%S')
df_transporte_filtrado['hora_completa'] = df_transporte_filtrado['hora_completa'].dt.strftime('%Y-%m-%dT%H:%M:%S')
df_vazao_final['Timestamp'] = df_vazao_final['Timestamp'].dt.strftime('%Y-%m-%dT%H:%M:%S')
df_vazao_final['data_hora_exata'] = df_vazao_final['data_hora_exata'].dt.strftime('%Y-%m-%dT%H:%M:%S')
df_dados_planta['Timestamp'] = df_dados_planta['Timestamp'].dt.strftime('%Y-%m-%dT%H:%M:%S')
df_totalizador['Timestamp'] = df_totalizador['Timestamp'].dt.strftime('%Y-%m-%dT%H:%M:%S')

#Mudar o nome da coluna para garantir compatibilidade com a tabela do supabase, depois vou eliminar esse trecho
df_totalizador.rename(columns={"Moinho_Massa Alimentada Moagem_(t)": "Retomada - TR02 - Balança"},inplace=True) #Moinho_Massa Alimentada Moagem_(t)

# 1. Converte colunas numéricas
colunas_numericas = ["Britagem_Massa Produzida Britagem_(t)"]
for col in colunas_numericas:
    df_dados_planta[col] = pd.to_numeric(df_dados_planta[col], errors="coerce")

# 2. Colunas de texto
colunas_texto = [
    "Britagem_Justificativa de NÂO atingir a massa_(txt)",
    "Moinho_Justificativa de NÂO atingir a massa_(txt)",
    "Moinho_Justificativa do Tempo operando com taxa a menor_(txt)"
]
for col in colunas_texto:
    df_dados_planta[col] = df_dados_planta[col].astype(str).replace({'nan': None, '': None})

# 3. Força todos os tipos para 'object'
for col in df_dados_planta.columns:
    df_dados_planta[col] = df_dados_planta[col].astype("object")

# 4. Substitui todos os NaN por None
df_dados_planta = df_dados_planta.where(pd.notnull(df_dados_planta), None)

# 5. Converte para dicionário e valida
dados = df_dados_planta.to_dict(orient='records')

import math
for i, linha in enumerate(dados):
    for chave, valor in linha.items():
        if isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)):
            raise ValueError(f"❌ Valor inválido detectado na linha {i}, coluna '{chave}': {valor}")

# =============================
# Enviar dados para o supabase
# =============================
# Supabase config
SUPABASE_URL = "https://bhjsqkraqjigjlrwueal.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJoanNxa3JhcWppZ2pscnd1ZWFsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTMxMjU3NTksImV4cCI6MjA2ODcwMTc1OX0.W10ta4ovNyFsVfc9u1OLA5-uzDzLETNbUAR_EOhMMvc"

#função para enviar os dados
def enviar_dados_para_supabase(df, table_name, url, key):
    supabase: Client = create_client(url, key)
    # Converte o DataFrame para dicionário
    dados = df.to_dict(orient='records')
    # Deleta todos os dados da tabela (usa condição para evitar erro)
    supabase.table(table_name).delete().neq("id", 0).execute()
    # Insere os novos dados
    resposta = supabase.table(table_name).insert(dados).execute()
    return resposta

# Chamada da função
resposta = enviar_dados_para_supabase(df_totalizador, 'alimentacao_moagem', SUPABASE_URL, SUPABASE_KEY)
resposta = enviar_dados_para_supabase(df_vazao_final, 'alimentacao_planta_media_movel', SUPABASE_URL, SUPABASE_KEY)
resposta = enviar_dados_para_supabase(df_transporte_filtrado, 'movimentacao_mina', SUPABASE_URL, SUPABASE_KEY)
resposta = enviar_dados_para_supabase(df_dados_planta, 'dados_planta', SUPABASE_URL, SUPABASE_KEY)